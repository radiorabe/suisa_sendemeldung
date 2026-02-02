"""SUISA Sendemeldung bugs SUISA with email once per month.

Fetches data on our playout history and formats them in a CSV file format
containing the data (like Track, Title and ISRC) requested by SUISA. Also
takes care of sending the report to SUISA via email for hands-off operations.
"""

from __future__ import annotations

from csv import reader, writer
from datetime import date, datetime, timedelta
from email.encoders import encode_base64
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formatdate
from io import BytesIO, StringIO
from pathlib import Path
from smtplib import SMTP
from string import Template
from typing import TYPE_CHECKING, TypeVar, cast

import click
import cridlib
import pytz
import typed_settings
from babel.dates import format_date
from dateutil.relativedelta import relativedelta
from iso3901 import ISRC
from openpyxl import Workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.styles import Border, Font, PatternFill, Side
from tqdm import tqdm
from typed_settings.cli_click import OptionGroupFactory
from typed_settings.exceptions import InvalidValueError

from suisa_sendemeldung.settings import FileFormat, IdentifierMode, OutputMode, Settings

from .acrclient import ACRClient

if TYPE_CHECKING:  # pragma: no cover
    from openpyxl.worksheet.worksheet import Worksheet


T = TypeVar("T")


def validate_arguments(settings: Settings) -> None:
    """Validate the arguments provided to the script.

    After this function we are sure that there are no conflicts in the arguments.

    Arguments:
    ---------
        settings: the Settings instance to validate

    Raises:
    ------
        InvalidValueErrors: if there are invalid argument combinations

    """
    msgs = []
    # xlsx cannot be printed to stdout
    if (
        settings.output == OutputMode.stdout
        and settings.file
        and settings.file.format == FileFormat.xlsx
    ):
        msgs.append("xlsx cannot be printed to stdout, please set --file-format to csv")
    # last_month is in conflict with start_date and end_date
    if settings.date.last_month and (settings.date.start or settings.date.end):
        msgs.append("argument --last-month not allowed with --date-start or --date-end")
    # exit if there are error messages
    if msgs:
        raise InvalidValueError(msgs)


def parse_date(settings: Settings) -> tuple[date, date]:
    """Parse date from args.

    Arguments:
    ---------
        settings: The settings provided to the script

    Returns:
    -------
        start_date: the start date of the requested interval
        end_date: the end date of the requested interval

    """
    # default values
    end_date: date = date.today()  # noqa: DTZ011
    start_date: date = end_date - timedelta(days=30)

    # date parsing logic
    if settings.date.last_month:
        today = date.today()  # noqa: DTZ011
        # get first of this month
        this_month = today.replace(day=1)
        # last day of last month = first day of this month - 1 day
        end_date = this_month - timedelta(days=1)
        start_date = end_date.replace(day=1)
    else:
        if settings.date.end:
            end_date = datetime.strptime(settings.date.end, "%Y-%m-%d").date()  # noqa: DTZ007
        if settings.date.start:
            start_date = datetime.strptime(settings.date.start, "%Y-%m-%d").date()  # noqa: DTZ007
    return start_date, end_date


def parse_filename(settings: Settings, start_date: date) -> str:
    """Parse filename from settings and start_date.

    Arguments:
    ---------
        settings: the settings provided to the script
        start_date: start of reporting period

    Returns:
    -------
        filename: the filename to use for the csv data

    """
    if settings.file.path:
        filename = settings.file.path
    # depending on date args either append the month or the start_date
    elif settings.date.last_month:
        date_part = f"{start_date.strftime('%Y')}_{start_date.strftime('%m')}"
        filename = f"{settings.station.name_short}_{date_part}.{settings.file.format}"
    else:
        filename = (
            f"{settings.station.name_short}_"
            f"{start_date.strftime('%Y-%m-%d')}.{settings.file.format}"
        )
    return filename


def check_duplicate(entry_a: dict, entry_b: dict) -> bool:
    """Check if two entries are duplicates by checking their acrid in all music items.

    Arguments:
    ---------
        entry_a: first entry
        entry_b: second entry

    Returns:
    -------
        True if the entries are duplicates, False otherwise

    """
    try:
        entry_a = entry_a["metadata"]["music"]
    except KeyError:
        entry_a = entry_a["metadata"]["custom_files"]
    try:
        entry_b = entry_b["metadata"]["music"]
    except KeyError:
        entry_b = entry_b["metadata"]["custom_files"]
    for music_a in entry_a:
        for music_b in entry_b:
            if music_a["acrid"] == music_b["acrid"]:
                return True
    return False


def merge_duplicates(data: list) -> list:
    """Merge consecutive entries into one if they are duplicates.

    Arguments:
    ---------
        data: The data provided by ACRClient

    Returns:
    -------
        data: The processed data

    """
    prev = data[0]
    mark = []
    for entry in data[1:]:
        if check_duplicate(prev, entry):
            prev["metadata"]["played_duration"] = (
                prev["metadata"]["played_duration"]
                + entry["metadata"]["played_duration"]
            )
            # mark entry for removal
            mark.append(entry)
        else:
            prev = entry
    # remove marked entries
    for entry in mark:
        data.remove(entry)
    return data


def funge_release_date(release_date: str = "") -> str:
    """Make a release_date from ACR conform to what seems to be the spec."""
    if len(release_date) == 10:  # noqa: PLR2004
        # we can make it look like what suisa has in their examples if it's the
        # right length
        try:
            return datetime.strptime(release_date, "%Y-%m-%d").strftime("%Y%m%d")  # noqa: DTZ007
        except ValueError:
            return ""
    # we discard other records since there is no way to convert records like a plain
    # year into dd/mm/yyyy properly without further guidance from whomever ingests
    # the data, in some cases this means we discard data that only contain a year
    # since they dont have that amount of precision.
    return ""


def get_artist(music: dict) -> str:
    """Get artist from a given dict.

    Arguments:
    ---------
        music: music dict from API

    Returns:
    -------
        artist: string representing the artist

    """
    artist = ""
    if music.get("artists") is not None:
        artists = music.get("artists")
        if isinstance(artists, list):
            artist = ", ".join([a.get("name") for a in artists])
        else:
            # Yet another 'wrong' entry in the database:
            # artists in custom_files was sometimes recorded as single value
            # @TODO also remove once way in the past? (2023-01-31)
            artist = cast("str", artists)
    elif music.get("artist") is not None:
        artist = cast("str", music.get("artist"))
    elif music.get("Artist") is not None:  # pragma: no cover
        # Uppercase is a hack needed for Jun 2021 since there is a 'wrong' entry
        # in the database. Going forward the record will be available as 'artist'
        # in lowercase.
        # @TODO remove once is waaaay in the past
        artist = cast("str", music.get("Artist"))
    return artist


def get_composer(music: dict) -> str:
    """Get composer from a given dict.

    Arguments:
    ---------
        music: music dict from API

    Returns:
    -------
        composer: string representing the composer

    """
    # composers is usually represented as a list of strings, e.g.:
    # 'contributors': {'composers': ['Alison Rachel Stewart', ...]}
    # if composers is not in the expected format, we just return an empty string
    composer = ""
    contributors = music.get("contributors")
    if contributors is not None:
        composers = contributors.get("composers")
        if composers is not None and isinstance(composers, list):
            composer = ", ".join(composers)
    return composer


def get_isrc(music: dict) -> str:
    """Get a valid ISRC from the music record or return an empty string."""
    isrc = ""
    if music.get("external_ids", {}).get("isrc"):
        isrc = music.get("external_ids", {}).get("isrc")
    elif music.get("isrc"):
        isrc = cast("str", music.get("isrc"))
    # was a list with a singular entry for a while back in 2021
    if isinstance(isrc, list):
        isrc = isrc[0]
    # some records contain the "ISRC" prefix that is described as legacy
    # in the ISRC handbook from IFPI.
    if isrc and isrc[:4] == "ISRC":
        isrc = isrc[4:]
    # take care of cases where the isrc is space delimited even though the
    # record is technically wrong but happens often enough to warrant this
    # hack.
    if isrc:
        isrc = isrc.replace(" ", "")

    if not ISRC.validate(isrc):
        isrc = ""
    return isrc


def get_csv(data: list, settings: Settings) -> str:
    """Create SUISA compatible csv data.

    Arguments:
    ---------
        data: To data to create csv from
        settings: The settings provided to the script

    Returns:
    -------
        csv: The converted data

    """
    station_name = settings.station.name
    header = [
        "Sender",
        "Titel des Musikwerks",
        "Name des Komponisten",
        "Interpret(en)",
        "Sendedatum",
        "Sendedauer",
        "Sendezeit",
        "ISRC",
        "Label",
        "Identifikationsnummer",
        "Eigenaufnahmen",
        "EAN / GTIN",
        "Albumtitel / Titel des Tonträgers",
        "Aufnahmedatum",
        "Aufnahmeland",
        "Erstveröffentlichungsdatum",
        "Katalog-Nummer / CD ID",
        "Werkverzeichnisangaben",
        "Bestellnummer",
        "Veröffentlichungsland",
        "Liveaufnahme",
    ]
    csv = StringIO()
    csv_writer = writer(csv, dialect="excel")
    csv_writer.writerow(header)

    for entry in tqdm(data, desc="preparing tracks for report"):
        metadata = entry.get("metadata")
        # parse timestamp
        timestamp = datetime.strptime(metadata.get("timestamp_local"), ACRClient.TS_FMT)  # noqa: DTZ007

        ts_date = timestamp.strftime("%Y-%m-%d")
        ts_time = timestamp.strftime("%H:%M:%S")
        hours, remainder = divmod(metadata.get("played_duration"), 60 * 60)
        minutes, seconds = divmod(remainder, 60)
        # required format of duration field: hh:mm:ss
        duration = f"{hours:02}:{minutes:02}:{seconds:02}"

        try:
            music = metadata.get("music")[0]
        except TypeError:
            music = metadata.get("custom_files")[0]
        title = music.get("title")

        artist = get_artist(music)
        composer = get_composer(music)

        works_composer = ", ".join(
            [
                c["name"]
                for c in [
                    item
                    for sublist in [w["creators"] for w in music.get("works", [])]
                    for item in sublist
                ]
                if c.get("role", "") in ["C", "Composer", "W", "Writer"]
            ],
        )
        if works_composer and (not composer or composer == artist):
            composer = works_composer

        isrc = get_isrc(music)
        label = music.get("label")

        # load some "best-effort" fields
        album = music.get("album", "")
        cd_id = ""
        # it's a dict if it's from the ACRCloud bucket, a string if from a custom bucket
        if isinstance(album, dict):
            cd_id = album.get("cd_id", "")
            album = album.get("name", "")
        upc = music.get("external_ids", {}).get("upc", "")
        release_date = funge_release_date(music.get("release_date", ""))

        local_id: str = ""
        # cridlib only supports timezone-aware datetime values, so we convert one
        timestamp_utc = pytz.utc.localize(
            datetime.strptime(metadata.get("timestamp_utc"), ACRClient.TS_FMT),  # noqa: DTZ007
        )
        # we include the acrid in our CRID so we know about the data's provenience
        # in case any questions about the data we delivered are asked
        acrid = music.get("acrid")

        if settings.crid_mode == IdentifierMode.cridlib:
            local_id = str(
                cridlib.get(timestamp=timestamp_utc, fragment=f"acrid={acrid}")
            )
        elif settings.crid_mode == IdentifierMode.local:
            local_id = f"{timestamp_utc.isoformat()}#acrid={acrid}"

        csv_writer.writerow(
            [
                station_name,
                title,
                composer,
                artist,
                ts_date,
                duration,
                ts_time,
                isrc,
                label,
                local_id,
                "nein",  # Eigenaufnahmen
                upc,
                album,
                "",  # Aufnahmedatum
                "",  # Aufnahmeland
                release_date,
                cd_id,
                "",  # Werkverzeichnisangaben
                "",  # Bestellnummer
                "",  # Veröffentlichungsland
                "",  # Liveaufnahme
            ],
        )
    return csv.getvalue()


def get_xlsx(data: list[dict], settings: Settings) -> BytesIO:
    """Create SUISA compatible xlsx data.

    Arguments:
    ---------
        data: The data to create xlsx from
        settings: The settings provided to the script

    Returns:
    -------
        xlsx: The converted data as BytesIO object

    """
    csv = get_csv(data, settings=settings)
    csv_reader = reader(StringIO(csv))

    xlsx = BytesIO()
    workbook: Workbook = Workbook()
    workbook.iso_dates = True
    if not workbook.active:  # pragma: no cover
        raise RuntimeError
    worksheet: Worksheet = workbook.active  # type: ignore[assignment]

    for row in csv_reader:
        worksheet.append(row)

    # the columns that should be styled as required (grey background)
    required_columns = [
        "Sender",
        "Titel des Musikwerks",
        "Name des Komponisten",
        "Interpret(en)",
        "Sendedatum",
        "Sendedauer",
        "Sendezeit",
        "ISRC",
        "Label",
        "Identifikationsnummer",
        "Eigenaufnahmen",
    ]
    subsidiary_columns = [
        "EAN/GTIN",
        "Albumtitel / Titel des Tonträgers",
        "Aufnahmedatum",
        "Aufnahmeland",
        "Erstveröffentlichungsdatum",
        "Katalog-Nummer / CD ID",
        "Werkverzeichnisangaben",
        "Bestellnummer",
    ]
    font = Font(name="Calibri", bold=True, size=12)
    side = Side(border_style="thick", color="000000")
    border = Border(top=side, left=side, right=side, bottom=side)
    required_fill = PatternFill("solid", bgColor="bfbfbf", fgColor="bfbfbf")
    subsdiary_fill = PatternFill("solid", bgColor="ebf1de", fgColor="ebf1de")
    for cell in worksheet[1]:  # xlsx is 1-indexed
        cell.font = font
        cell.border = border
        if cell.value in required_columns:
            cell.fill = required_fill
        elif cell.value in subsidiary_columns:
            cell.fill = subsdiary_fill

    # Try to approximate the required width by finding the longest values per column
    dims: dict[str, int] = {}
    calc_row: tuple[Cell | MergedCell, ...]
    for calc_row in worksheet.rows:
        for cell in calc_row:
            if isinstance(cell, Cell) and cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value))),
                )
    # apply estimated width to each column
    padding = 3
    for col, value in dims.items():
        worksheet.column_dimensions[col].width = value + padding

    reformat_start_date_in_xlsx(worksheet)

    workbook.save(xlsx)
    return xlsx


def reformat_start_date_in_xlsx(worksheet: Worksheet) -> None:
    """Set date number formatting on relevant columns."""
    for idx, row in enumerate(worksheet.rows):
        # skip first row
        if idx < 1:
            continue

        # turn the str from the CSV into a real datetime.datetime in Sendedatum column
        row[4].value = datetime.strptime(  # noqa: DTZ007
            f"{row[4].value} {row[6].value}", "%Y-%m-%d %H:%M:%S"
        ).date()  # pyright: ignore[reportAttributeAccessIssue]
        # adjust the formatting
        row[4].number_format = "dd.mm.yyyy"

        # same thing for date fields "Aufnahmedatum" and "Erstveröffentlichungsdatum"
        for col_idx in [13, 15]:
            row[col_idx].value = (
                datetime.strptime(  # noqa: DTZ007
                    str(row[col_idx].value), "%Y%m%d"
                ).date()
                if row[col_idx].value
                else None
            )  # pyright: ignore[reportAttributeAccessIssue]
            row[col_idx].number_format = "dd.mm.yyyy"


def write_csv(filename: str, csv: BytesIO | str) -> None:  # pragma: no cover
    """Write contents of `csv` to file.

    Arguments:
    ---------
        filename: The file to write to.
        csv: The data to write to `filename`.

    """
    with Path(filename).open("w", encoding="utf-8") as csvfile:
        csvfile.write(str(csv))


def write_xlsx(filename: str, xlsx: BytesIO) -> None:  # pragma: no cover
    """Write contents of `xlsx` to file.

    Arguments:
    ---------
        filename: The file to write to.
        xlsx: The data to write to `filename`.

    """
    with Path(filename).open("wb") as xlsxfile:
        xlsxfile.write(xlsx.getvalue())


def get_email_attachment(filename: str, filetype: str, data: BytesIO | str) -> MIMEBase:
    """Create attachment based on required filetype and data.

    Arguments:
    ---------
        filename: The filename of the attachment
        filetype: The filetype of the attachment
        data: The attachment data

    """
    maintype = "application"
    subtype = "vnd.ms-excel"
    if filetype == "csv":
        maintype = "text"
        subtype = "csv"
        part = MIMEBase("text", "csv")

    payload = (
        data.getvalue() if isinstance(data, BytesIO) else str(data).encode("utf-8")
    )

    part = MIMEBase(maintype, subtype)
    part.set_payload(payload)
    encode_base64(part)
    part.add_header(
        "Content-Disposition", f"attachment; filename={Path(filename).name}"
    )
    return part


def create_message(  # noqa: PLR0913
    sender: str,
    recipient: str,
    subject: str,
    text: str,
    filename: str,
    filetype: str,
    data: BytesIO | str,
    cc: str | None = None,
    bcc: str | None = None,
) -> MIMEMultipart:
    """Create email message.

    Arguments:
    ---------
        sender: The sender of the email. Login will be made with this user.
        recipient: The recipient of the email. Can be a list.
        subject: The subject of the email.
        text: The body of the email.
        filename: The filename of the attachment
        filetype: The filetype of the attachment
        data: The attachment data.
        cc: cc recipient
        bcc: bcc recipient

    """
    msg = MIMEMultipart()
    msg["From"] = sender
    msg["To"] = recipient
    if cc:
        msg["Cc"] = cc
    if bcc:
        msg["Bcc"] = bcc
    msg["Date"] = formatdate(localtime=True)
    msg["Subject"] = subject
    # set body
    msg.attach(MIMEText(text))
    msg.attach(get_email_attachment(filename, filetype, data))

    return msg


def send_message(
    msg: MIMEMultipart,
    server: str = "127.0.0.1",
    port: int = 587,
    login: str | None = None,
    password: str | None = None,
) -> None:
    """Send email.

    Arguments:
    ---------
        msg: The message to send (an email.messag.Message object)
        server: The SMTP server to use to send the email.
        port: The port of the SMTP server.
        login: The username for `sender`@`server`.
        password: The password for `sender`@`server`.

    """
    with SMTP(host=server, port=port) as smtp:
        smtp.starttls()
        if password:
            if login:
                smtp.login(login, password)
            else:
                smtp.login(msg["From"], password)
        smtp.send_message(msg)


def main(settings: Settings) -> None:  # pragma: no cover
    """ACRCloud client for SUISA reporting @ RaBe."""
    validate_arguments(settings)

    start_date, end_date = parse_date(settings)
    filename = parse_filename(settings, start_date)

    client = ACRClient(bearer_token=str(settings.acr.bearer_token))
    data = client.get_interval_data(
        settings.acr.project_id,
        str(settings.acr.stream_id),
        start_date,
        end_date,
        timezone=settings.l10n.timezone,
    )
    data = merge_duplicates(data)
    if settings.file.format == FileFormat.xlsx:
        data = get_xlsx(data, settings=settings)
    elif settings.file.format == FileFormat.csv:
        data = get_csv(data, settings=settings)

    if settings.output == OutputMode.email:
        email_subject = Template(settings.email.subject).substitute(
            {
                "station_name": settings.station.name,
                "year": format_date(
                    start_date, format="yyyy", locale=settings.l10n.locale
                ),
                "month": format_date(
                    start_date, format="MM", locale=settings.l10n.locale
                ),
            },
        )
        # generate body
        text = Template(settings.email.text).substitute(
            {
                "station_name": settings.station.name,
                "month": format_date(
                    start_date, format="MMMM", locale=settings.l10n.locale
                ),
                "year": format_date(
                    start_date, format="yyyy", locale=settings.l10n.locale
                ),
                "previous_year": format_date(
                    start_date - timedelta(days=365),
                    format="yyyy",
                    locale=settings.l10n.locale,
                ),
                "in_three_months": format_date(
                    datetime.now() + relativedelta(months=+3),  # noqa: DTZ005
                    format="long",
                    locale=settings.l10n.locale,
                ),
                "responsible_email": settings.email.responsible_email,
                "email_footer": settings.email.footer,
            },
        )
        msg = create_message(
            settings.email.sender,
            settings.email.to,
            email_subject,
            text,
            filename,
            settings.file.format,
            data,  # pyright: ignore[reportArgumentType]
            cc=settings.email.cc,
            bcc=settings.email.bcc,
        )
        send_message(
            msg,
            server=settings.email.server,
            port=settings.email.port,
            login=settings.email.username,
            password=settings.email.password,
        )

    elif settings.output == OutputMode.file and settings.file.format == FileFormat.xlsx:
        write_xlsx(filename, data)  # pyright: ignore[reportArgumentType]
    elif settings.output == OutputMode.file and settings.file.format == FileFormat.csv:
        write_csv(filename, data)  # pyright: ignore[reportArgumentType]
    elif (
        settings.output == OutputMode.stdout and settings.file.format == FileFormat.csv
    ):
        print(data)  # noqa: T201


@click.command()
@typed_settings.click_options(
    Settings,
    loaders=typed_settings.default_loaders(
        "sendemeldung",
        [
            "/etc/suisa_sendemeldung.toml",
            "suisa_sendemeldung.toml",
        ],
    ),
    decorator_factory=OptionGroupFactory(),
    show_envvars_in_help=True,
)
def cli(settings: Settings) -> None:  # pragma: no cover
    """SUISA Sendemeldung.

    Create and send playout reports to SUISA.

    The reports are based on data from ACRCloud.
    """
    main(settings)


if __name__ == "__main__":  # pragma: no cover
    cli()
