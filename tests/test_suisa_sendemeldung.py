"""Test the suisa_sendemeldung.suisa_sendemeldung module."""

from datetime import date, datetime, timezone
from email.message import Message
from io import BytesIO
from typing import TYPE_CHECKING
from unittest.mock import call, patch

import pytest
from click.testing import CliRunner
from freezegun import freeze_time
from openpyxl import Workbook, load_workbook
from typed_settings.exceptions import InvalidValueError

from suisa_sendemeldung import suisa_sendemeldung
from suisa_sendemeldung.settings import (
    ACR,
    FileFormat,
    FileSettings,
    OutputMode,
    RangeSettings,
    Settings,
    StationSettings,
)

if TYPE_CHECKING:  # pragma: no cover
    from openpyxl.worksheet.worksheet import Worksheet


def test_validate_arguments():
    """Test validate_arguments."""

    with pytest.raises(ValueError) as excinfo:  # noqa: PT011
        Settings(
            acr=ACR(
                bearer_token="short",
                project_id=123456,
                stream_id="stream123",
            )
        )
    assert "Length of 'bearer_token' must be >= 32: 5" in str(excinfo.value)

    with pytest.raises(ValueError) as excinfo:  # noqa: PT011
        Settings(
            acr=ACR(
                bearer_token="_" * 32,
                project_id=-1,  # invalid project id
                stream_id="stream123",
            )
        )
    assert "'project_id' must be >= 0: -1" in str(excinfo.value)

    with pytest.raises(ValueError) as excinfo:  # noqa: PT011
        Settings(
            acr=ACR(
                bearer_token="_" * 32,
                project_id=1,
                stream_id="0",  # invalid stream id
            )
        )
    assert "Length of 'stream_id' must be >= 9: 1" in str(excinfo.value)

    settings = Settings()
    # last_month is in conflict with start_date and end_date
    settings.date.last_month = True
    settings.date.start = date(1993, 3, 1).strftime("%Y-%m-%d")

    with pytest.raises(InvalidValueError) as excinfo:
        suisa_sendemeldung.validate_arguments(settings)
    assert "argument --last-month not allowed with --date-start or --date-end" in str(
        excinfo.value
    )

    settings = Settings()
    settings.output = OutputMode.stdout
    settings.file.format = FileFormat.xlsx
    with pytest.raises(InvalidValueError) as excinfo:
        suisa_sendemeldung.validate_arguments(settings)
    assert "xlsx cannot be printed to stdout, please set --file-format to csv" in str(
        excinfo.value
    )


def test_parse_date():
    """Test parse_date."""

    # default, "last_month" case
    with freeze_time("1996-04-01"):
        settings = Settings()
        settings.date.last_month = True
        (start_date, end_date) = suisa_sendemeldung.parse_date(settings)
    assert start_date == date(1996, 3, 1)
    assert end_date == date(1996, 3, 31)

    # if no start_date was set, use the current data as start_date
    with freeze_time("1996-03-01"):
        settings = Settings()
        settings.date.last_month = False
        settings.date.start = ""
        settings.date.end = "1996-03-31"
        (start_date, end_date) = suisa_sendemeldung.parse_date(settings)
    assert start_date == date(1996, 1, 31)
    assert end_date == date(1996, 3, 31)

    # if no end_date was set, default to today
    with freeze_time("1996-03-31"):
        settings = Settings()
        settings.date.last_month = False
        settings.date.start = ""
        settings.date.end = ""
        (start_date, end_date) = suisa_sendemeldung.parse_date(settings)
    assert start_date == date(1996, 3, 1)
    assert end_date == date(1996, 3, 31)

    # only specify start_date, selects up to today
    with freeze_time("1996-04-04"):
        settings = Settings()
        settings.date.last_month = False
        settings.date.start = "1996-03-01"
        settings.date.end = ""
        (start_date, end_date) = suisa_sendemeldung.parse_date(settings)
    assert start_date == date(1996, 3, 1)
    assert end_date == date(1996, 4, 4)


def test_parse_filename():
    """Test parse_filename."""

    # pass filename from cli
    settings = Settings(
        file=FileSettings(path="/foo/bar"),
        station=StationSettings(name_short="test"),
    )
    filename = suisa_sendemeldung.parse_filename(settings, datetime.now())
    assert filename == "/foo/bar"

    # last_month mode
    settings = Settings(
        date=RangeSettings(last_month=True),
        file=FileSettings(format=FileFormat.xlsx),
        station=StationSettings(name_short="test"),
    )
    with freeze_time("1996-03-01"):
        filename = suisa_sendemeldung.parse_filename(settings, datetime.now())
    assert filename == "test_1996_03.xlsx"

    # start date mode
    settings = Settings(
        date=RangeSettings(last_month=False, start="1996-03-01"),
        file=FileSettings(format=FileFormat.xlsx),
        station=StationSettings(name_short="test"),
    )
    with freeze_time("1996-03-01"):
        filename = suisa_sendemeldung.parse_filename(settings, datetime.now())
    assert filename == "test_1996-03-01.xlsx"


def test_check_duplicate():
    """Test check_duplicates."""

    # both records are music and not duplicate
    entry_a = {"metadata": {"music": [{"acrid": "123456789"}]}}
    entry_b = {"metadata": {"music": [{"acrid": "987654321"}]}}
    assert not suisa_sendemeldung.check_duplicate(entry_a, entry_b)

    # both records are music and duplicate
    entry_a = {"metadata": {"music": [{"acrid": "123456789"}]}}
    entry_b = {"metadata": {"music": [{"acrid": "123456789"}]}}
    assert suisa_sendemeldung.check_duplicate(entry_a, entry_b)

    # first record is custom and not duplicate
    entry_a = {"metadata": {"custom_files": [{"acrid": "123456789"}]}}
    entry_b = {"metadata": {"music": [{"acrid": "987654321"}]}}
    assert not suisa_sendemeldung.check_duplicate(entry_a, entry_b)

    # first record is custom and duplicate
    entry_a = {"metadata": {"custom_files": [{"acrid": "123456789"}]}}
    entry_b = {"metadata": {"music": [{"acrid": "123456789"}]}}
    assert suisa_sendemeldung.check_duplicate(entry_a, entry_b)

    # second record is custom and not duplicate
    entry_a = {"metadata": {"music": [{"acrid": "123456789"}]}}
    entry_b = {"metadata": {"custom_files": [{"acrid": "987654321"}]}}
    assert not suisa_sendemeldung.check_duplicate(entry_a, entry_b)

    # second record is custom and duplicate
    entry_a = {"metadata": {"music": [{"acrid": "123456789"}]}}
    entry_b = {"metadata": {"custom_files": [{"acrid": "123456789"}]}}
    assert suisa_sendemeldung.check_duplicate(entry_a, entry_b)

    # both records are custom and not duplicate
    entry_a = {"metadata": {"custom_files": [{"acrid": "123456789"}]}}
    entry_b = {"metadata": {"custom_files": [{"acrid": "987654321"}]}}
    assert not suisa_sendemeldung.check_duplicate(entry_a, entry_b)

    # both records are custom and duplicate
    entry_a = {"metadata": {"custom_files": [{"acrid": "123456789"}]}}
    entry_b = {"metadata": {"custom_files": [{"acrid": "123456789"}]}}
    assert suisa_sendemeldung.check_duplicate(entry_a, entry_b)


def test_merge_duplicates():
    """Test merge_duplicates."""

    # check if record_1 reduced to one record and durations are added together
    record_1 = {"metadata": {"music": [{"acrid": "123456789"}], "played_duration": 10}}
    record_2 = {"metadata": {"music": [{"acrid": "987654321"}], "played_duration": 10}}
    raw_records = [record_1, record_1, record_2]
    results = suisa_sendemeldung.merge_duplicates(raw_records)
    assert len(results) == 2  # noqa: PLR2004
    assert results[0]["metadata"]["played_duration"] == 20  # noqa: PLR2004


@pytest.mark.parametrize(
    ("test_date", "expected"),
    [
        ("0000-00-00", ""),
    ],
)
def test_funge_release_date(test_date, expected):
    """Test funge_release_date."""

    results = suisa_sendemeldung.funge_release_date(test_date)
    assert results == expected


@patch("cridlib.get")
def test_get_csv(mock_cridlib_get, snapshot, settings):
    """Test get_csv."""
    mock_cridlib_get.return_value = "crid://rabe.ch/v1/test"

    # empty data
    data = []
    csv = suisa_sendemeldung.get_csv(data, settings=settings)
    assert csv == snapshot
    mock_cridlib_get.assert_not_called()

    # bunch of data
    mock_cridlib_get.reset_mock()
    data = [
        {
            "metadata": {
                "timestamp_local": "1993-03-01 13:12:00",
                "timestamp_utc": "1993-03-01 13:12:00",
                "played_duration": 60,
                "music": [{"title": "Uhrenvergleich", "acrid": "a1"}],
            },
        },
        {
            "metadata": {
                "timestamp_local": "1993-03-01 13:37:00",
                "timestamp_utc": "1993-03-01 13:37:00",
                "played_duration": 60,
                "custom_files": [
                    {
                        "acrid": "a2",
                        "title": "Meme Dub",
                        "artist": "Da Gang",
                        "album": "album, but string",
                        "contributors": {
                            "composers": [
                                "Da Composah",
                            ],
                        },
                        "release_date": "2023",
                        "external_ids": {"isrc": "DEZ650710376"},
                    },
                ],
            },
        },
        {
            "metadata": {
                "timestamp_local": "1993-03-01 16:20:00",
                "timestamp_utc": "1993-03-01 16:20:00",
                "played_duration": 60,
                "music": [
                    {
                        "acrid": "a3",
                        "title": "Bubbles",
                        "album": {
                            "name": "Da Alboom",
                        },
                        "contributors": {"composers": None},
                        "release_date": "2022-12-13",
                        "artists": [
                            {
                                "name": "Mary's Surprise Act",
                            },
                            {
                                "name": "Climmy Jiff",
                            },
                        ],
                        "isrc": "DEZ650710376",
                        "label": "Jane Records",
                        "external_ids": {
                            "upc": "greedy-capitalist-number",
                        },
                    },
                ],
            },
        },
        {
            "metadata": {
                "timestamp_local": "1993-03-01 17:17:17",
                "timestamp_utc": "1993-03-01 17:17:17",
                "played_duration": 60,
                "custom_files": [
                    {
                        "acrid": "a4",
                        "artists": "Artists as string not list",
                    },
                ],
            },
        },
        {
            "metadata": {
                "timestamp_local": "1993-03-01 18:18:18",
                "timestamp_utc": "1993-03-01 18:18:18",
                "played_duration": 71337,
                "music": [{"title": "Long Playing", "acrid": "a5"}],
            },
        },
        {
            "metadata": {
                "timestamp_local": "1993-03-01 18:18:18",
                "timestamp_utc": "1993-03-01 18:18:18",
                "played_duration": 71337,
                "music": [
                    {
                        "title": "composer in works",
                        "acrid": "a6",
                        "works": [{"creators": [{"name": "Worker", "role": "W"}]}],
                    },
                ],
            },
        },
        {
            "metadata": {
                "timestamp_local": "1993-03-01 18:18:18",
                "timestamp_utc": "1993-03-01 18:18:18",
                "played_duration": 71337,
                "music": [
                    {
                        "title": "composer better in works",
                        "artists": [{"name": "same"}],
                        "contributors": {
                            "composers": ["same"],
                        },
                        "acrid": "a6",
                        "works": [{"creators": [{"name": "composer", "role": "C"}]}],
                    },
                ],
            },
        },
    ]
    csv = suisa_sendemeldung.get_csv(data, settings=settings)
    assert csv == snapshot
    mock_cridlib_get.assert_has_calls(
        [
            call(
                timestamp=datetime(1993, 3, 1, 13, 12, tzinfo=timezone.utc),
                fragment="acrid=a1",
            ),
            call(
                timestamp=datetime(1993, 3, 1, 13, 37, tzinfo=timezone.utc),
                fragment="acrid=a2",
            ),
            call(
                timestamp=datetime(1993, 3, 1, 16, 20, tzinfo=timezone.utc),
                fragment="acrid=a3",
            ),
            call(
                timestamp=datetime(1993, 3, 1, 17, 17, 17, tzinfo=timezone.utc),
                fragment="acrid=a4",
            ),
            call(
                timestamp=datetime(1993, 3, 1, 18, 18, 18, tzinfo=timezone.utc),
                fragment="acrid=a5",
            ),
        ],
    )

    # no cridib
    mock_cridlib_get.reset_mock()
    settings.crid_mode = "local"
    csv = suisa_sendemeldung.get_csv(data, settings=settings)
    assert csv == snapshot
    mock_cridlib_get.assert_not_called()


def test_get_xlsx(snapshot, settings):
    """Test get_xlsx."""

    # empty data
    data = []
    xlsx = suisa_sendemeldung.get_xlsx(data, settings=settings)
    workbook = load_workbook(xlsx)
    worksheet = workbook.active
    assert list(worksheet.values) == snapshot  # pyright: ignore[reportOptionalMemberAccess]
    assert worksheet.column_dimensions == snapshot  # pyright: ignore[reportOptionalMemberAccess]


def test_reformat_start_date_in_xlsx():
    """Test that reformat_start_date_in_xlsx reformats the start date column."""
    workbook: Workbook = Workbook()
    if not workbook.active:  # pragma: no cover
        raise RuntimeError
    worksheet: Worksheet = workbook.active  # type: ignore[assignment]
    worksheet.append([])
    worksheet.append(
        [
            "",
            "",
            "",
            "",
            "2025-01-01",  # Sendedatum
            "",
            "01:01:01",  # Sendezeit
            "",
            "",
            "",
            "",
            "",
            "",
            "20250101",  # Aufnahmedatum
            "",
            "20250101",  # Erstveröffentlichungsdatum
        ]
    )
    suisa_sendemeldung.reformat_start_date_in_xlsx(worksheet)
    row = list(worksheet.rows)[1]
    assert row[4].value == datetime(2025, 1, 1).date()
    assert row[4].number_format == "dd.mm.yyyy"
    assert row[13].value == datetime(2025, 1, 1).date()
    assert row[13].number_format == "dd.mm.yyyy"
    assert row[15].value == datetime(2025, 1, 1).date()
    assert row[15].number_format == "dd.mm.yyyy"


def test_get_email_attachment():
    """Test get_email_attachment."""
    filename = "test.xlsx"
    filetype = "xlsx"
    data = BytesIO()
    part = suisa_sendemeldung.get_email_attachment(filename, filetype, data)
    assert part.get_filename() == "test.xlsx"
    assert part.get_content_type() == "application/vnd.ms-excel"

    filename = "test.csv"
    filetype = "csv"
    data = "data"
    part = suisa_sendemeldung.get_email_attachment(filename, filetype, data)
    assert part.get_filename() == "test.csv"
    assert part.get_content_type() == "text/csv"


def test_create_message():
    """Test create_message."""
    sender = "from@example.org"
    recipient = "long-arm-of-music-industry@example.com"
    subject = "subject"
    text = "text"
    filename = "/tmp/filename"
    filetype = "csv"
    data = "data"
    carbon_copy = "cc@example.org"
    bcc = "bcc@example.org"

    msg = suisa_sendemeldung.create_message(
        sender,
        recipient,
        subject,
        text,
        filename,
        filetype,
        data,
        carbon_copy,
        bcc,
    )
    assert msg.get("From") == sender
    assert msg.get("To") == recipient
    assert msg.get("Subject") == subject
    assert msg.get("Cc") == carbon_copy
    assert msg.get("Bcc") == bcc


def test_send_message():
    """Test send_message."""
    msg = Message()

    # no auth
    with patch("suisa_sendemeldung.suisa_sendemeldung.SMTP", autospec=True) as mock:
        suisa_sendemeldung.send_message(msg)  # pyright: ignore[reportArgumentType]
        mock.assert_called_once_with("127.0.0.1", 587)
        ctx = mock.return_value.__enter__.return_value
        ctx.starttls.assert_called_once()
        ctx.send_message.assert_called_once_with(msg)

    # auth, user provided login
    with patch("suisa_sendemeldung.suisa_sendemeldung.SMTP", autospec=True) as mock:
        suisa_sendemeldung.send_message(msg, "127.0.0.1", 587, "user", "password")  # pyright: ignore[reportArgumentType]
        mock.assert_called_once_with("127.0.0.1", 587)
        ctx = mock.return_value.__enter__.return_value
        ctx.starttls.assert_called_once()
        ctx.login.assert_called_once_with("user", "password")

    # auth, user from msg
    with patch("suisa_sendemeldung.suisa_sendemeldung.SMTP", autospec=True) as mock:
        msg.add_header("From", "test@example.org")
        suisa_sendemeldung.send_message(msg, "127.0.0.1", 587, None, "password")  # pyright: ignore[reportArgumentType]
        mock.assert_called_once_with("127.0.0.1", 587)
        ctx = mock.return_value.__enter__.return_value
        ctx.starttls.assert_called_once()
        ctx.login.assert_called_once_with("test@example.org", "password")


@pytest.mark.parametrize(
    ("test_music", "expected"),
    [
        ({"external_ids": {"isrc": "DEZ650710376"}}, "DEZ650710376"),
        ({"external_ids": {"isrc": ["DEZ650710376"]}}, "DEZ650710376"),
        ({"external_ids": {"isrc": "DE Z65 07 10376"}}, "DEZ650710376"),
        ({"external_ids": {"isrc": "ISRCDEZ650710376"}}, "DEZ650710376"),
        ({"external_ids": {"isrc": "123456789-1"}}, ""),
        ({"isrc": "DEZ650710376"}, "DEZ650710376"),
        ({"isrc": ["DEZ650710376"]}, "DEZ650710376"),
        ({"isrc": "ISRCDEZ650710376"}, "DEZ650710376"),
        ({"isrc": "DE Z65 07 10376"}, "DEZ650710376"),
        ({"isrc": "123456789-1"}, ""),
    ],
)
def test_get_isrc(test_music, expected):
    """Test get_isrc."""

    isrc = suisa_sendemeldung.get_isrc(test_music)
    assert isrc == expected


def test_cli_help(snapshot):
    """Snapshot test cli output."""
    runner = CliRunner()
    # Invoke the command with the --help option
    result = runner.invoke(suisa_sendemeldung.cli, ["--help"])
    assert result.output == snapshot
